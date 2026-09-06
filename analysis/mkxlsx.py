# -*- coding: utf-8 -*-
"""Build the fill-in workbook for the nine labels that need checking."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

D = json.load(open('cases9.json'))
F = 'Arial'
HEAD = PatternFill('solid', fgColor='1F3864')
FILLIN = PatternFill('solid', fgColor='FFFF00')      # cells the user edits
T1 = PatternFill('solid', fgColor='F8CBAD')          # tier 1, highest suspicion
T2 = PatternFill('solid', fgColor='FFF2CC')
T3 = PatternFill('solid', fgColor='E2EFDA')
THIN = Border(*[Side('thin', color='BFBFBF')] * 4)
WRAP = Alignment(wrap_text=True, vertical='top')
TIERFILL = {'一档': T1, '二档': T2, '三档': T3}

wb = Workbook()

# ---------------------------------------------------------------- 说明
ws = wb.active; ws.title = '说明'
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 96
lines = [
 ('T', '待核标签清单：最终标签为阳性、但导出报告文本中找不到依据的 9 例', ''),
 ('', '', ''),
 ('H', '这份清单是怎么来的', ''),
 ('P', '来源', '对全部 740 个索引检查次（患者×模态）做双向核对：判阳性的，文本里有没有诊断名或该模态征象；'
        '判阴性的，文本里有没有点名诊断。脚本 analysis/labelaudit.py，可复跑。'),
 ('P', '结果', '判阳性而文本无支持 9 例；反方向 0 例。唯一一例判阴性而文本含诊断名者，原文写的是'
        '「未探及明显"肠旋转不良"征象」，属明确否定，判阴性正确。'),
 ('P', '重要限制', '词面正则看不见「描述了征象但没点名」的写法，这正是 Online Resource 1 G 节列的失败模式 (ii)。'
        '所以被标记出来的是待读的病例，不是已证实的错误——二档四例读下来很可能裁定是对的。'),
 ('', '', ''),
 ('H', '三个分档', ''),
 ('P', '一档 高度可疑', '该模态全部术前报告中没有任何旋转相关内容。3 例。建议优先核这三例。'),
 ('P', '二档 可能是对的', '报告用形态学语言描述了征象，只是没用「漩涡」这类词，或提及了正文判据未收录的征象。4 例。'),
 ('P', '三档 已有解释', '一例结论仅为交叉引用、被引报告不在导出数据中；一例漩涡征记录在更早一次同模态检查上，'
        'Online Resource 1 H 节已披露。2 例。'),
 ('', '', ''),
 ('H', '请你填哪几列', ''),
 ('P', '「待核清单」工作表', '黄色底的三列由你填写：J 列「核对结论」用下拉选择；K 列「实际依据」写清当时判阳性的根据；'
        'L 列「备注」可选。其余列请勿改动，「影响测算」工作表按 J 列自动重算。'),
 ('P', '填写示例', '假设第 4 行（患者 4331826）你查病历后确认超声/CT 当时确实没报出旋转不良：'
        'J 列选「确为错标（改为阴性）」，K 列填「病历核对：术前 CT 报告未提及旋转不良，诊断由术中所见确立」，'
        'L 列填「2026-09-06 由 XXX 核」。'),
 ('', '', ''),
 ('H', '核完之后', ''),
 ('P', '下一步', '把填好的表发回给我。凡标记为「确为错标」的，我会改动裁定矩阵对应单元格、重跑全部分析，'
        '并同步摘要、正文、四张表、三张图与全部模型。标记为「裁定正确」的，我会在 Methods 里补一句说明'
        '形态学描述经裁定后同样计为阳性，使定义与标签一致。'),
 ('P', '数据出处', '报告原文取自 全部肠旋转不良数据.xlsx；最终标签取自 诊断效能_逐患者矩阵_当前版v3.xlsx。'),
]
r = 1
for kind, b, c in lines:
    if kind == 'T':
        ws.cell(r, 2, b).font = Font(F, size=14, bold=True, color='1F3864')
    elif kind == 'H':
        ws.cell(r, 2, b).font = Font(F, size=11, bold=True, color='1F3864')
    elif kind == 'P':
        ws.cell(r, 2, b).font = Font(F, size=10, bold=True)
        cc = ws.cell(r, 3, c); cc.font = Font(F, size=10); cc.alignment = WRAP
        ws.row_dimensions[r].height = 15 + 13 * (len(c) // 60)
    r += 1

# ---------------------------------------------------------------- 待核清单
ws = wb.create_sheet('待核清单')
hdr = ['序号', '患者编号', '模态', '术前该模态报告数', '分档', '标记原因',
       '索引检查次结论原文', '现最终标签', '', '核对结论', '实际依据', '备注']
for j, h in enumerate(hdr, 1):
    c = ws.cell(1, j, h)
    c.font = Font(F, size=10, bold=True, color='FFFFFF')
    c.fill = HEAD; c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    c.border = THIN
ws.row_dimensions[1].height = 30
for i, x in enumerate(D['rows'], start=2):
    vals = [x['no'], x['pid'], x['mod'], x['n'], x['tier'], x['note'],
            x['concl'], '阳性', '', '', '', '']
    for j, v in enumerate(vals, 1):
        c = ws.cell(i, j, v)
        c.font = Font(F, size=10); c.alignment = WRAP; c.border = THIN
        if j == 5:
            c.fill = TIERFILL[x['tier'][:2]]
        if j in (10, 11, 12):
            c.fill = FILLIN
    ws.row_dimensions[i].height = 58
for col, w in zip('ABCDEFGHIJKL', [5, 11, 6, 9, 15, 34, 46, 9, 2, 20, 34, 16]):
    ws.column_dimensions[col].width = w
dv = DataValidation(type='list', allow_blank=True, showDropDown=False,
                    formula1='"裁定正确（保持阳性）,确为错标（改为阴性）,导出数据漏了报告,尚无法判定"')
ws.add_data_validation(dv); dv.add(f'J2:J{len(D["rows"])+1}')
ws.freeze_panes = 'A2'

# ---------------------------------------------------------------- 报告全文
ws = wb.create_sheet('报告全文')
hdr = ['序号', '患者编号', '模态', '距手术天数', '检查名称', '检查所见', '检查结论']
for j, h in enumerate(hdr, 1):
    c = ws.cell(1, j, h)
    c.font = Font(F, size=10, bold=True, color='FFFFFF'); c.fill = HEAD
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center'); c.border = THIN
for i, x in enumerate(D['reps'], start=2):
    for j, v in enumerate([x['no'], x['pid'], x['mod'], x['gap'], x['name'], x['find'], x['concl']], 1):
        c = ws.cell(i, j, v)
        c.font = Font(F, size=9); c.alignment = WRAP; c.border = THIN
    ws.row_dimensions[i].height = 96
for col, w in zip('ABCDEFG', [5, 11, 6, 11, 30, 78, 52]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'

# ---------------------------------------------------------------- 影响测算
ws = wb.create_sheet('影响测算')
ws.cell(1, 1, '更正的影响：按「待核清单」J 列自动重算').font = Font(F, size=12, bold=True, color='1F3864')
ws.cell(2, 1, '仅统计标记为「确为错标（改为阴性）」的病例。Wilson 95% 置信区间按公式实时计算。').font = Font(F, size=9, italic=True)
hdr = ['模态', '现分子', '分母', '现检出率', '现 95% CI 下限', '现 95% CI 上限',
       '判为错标数', '更正后分子', '更正后检出率', '更正后 CI 下限', '更正后 CI 上限']
for j, h in enumerate(hdr, 1):
    c = ws.cell(4, j, h)
    c.font = Font(F, size=10, bold=True, color='FFFFFF'); c.fill = HEAD
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center'); c.border = THIN
ws.row_dimensions[4].height = 30
n = len(D['rows']) + 1
for i, (mod, k, tot) in enumerate([('UGI', 237, 301), ('CT', 171, 320), ('US', 65, 119)], start=5):
    ws.cell(i, 1, mod)
    ws.cell(i, 2, k)
    ws.cell(i, 3, tot)
    ws.cell(i, 4, f'=B{i}/C{i}')
    # Wilson interval, z = 1.96
    for col, sgn in (('E', '-'), ('F', '+')):
        ws.cell(i, 5 if col == 'E' else 6,
                f'=((B{i}/C{i}+1.96^2/(2*C{i})){sgn}1.96*SQRT((B{i}/C{i})*(1-B{i}/C{i})/C{i}'
                f'+1.96^2/(4*C{i}^2)))/(1+1.96^2/C{i})')
    ws.cell(i, 7, f'=COUNTIFS(待核清单!$C$2:$C${n},A{i},待核清单!$J$2:$J${n},"确为错标（改为阴性）")')
    ws.cell(i, 8, f'=B{i}-G{i}')
    ws.cell(i, 9, f'=H{i}/C{i}')
    for col, sgn in (('J', '-'), ('K', '+')):
        ws.cell(i, 10 if col == 'J' else 11,
                f'=((H{i}/C{i}+1.96^2/(2*C{i})){sgn}1.96*SQRT((H{i}/C{i})*(1-H{i}/C{i})/C{i}'
                f'+1.96^2/(4*C{i}^2)))/(1+1.96^2/C{i})')
    for j in range(1, 12):
        c = ws.cell(i, j); c.font = Font(F, size=10); c.border = THIN
        if j in (4, 5, 6, 9, 10, 11):
            c.number_format = '0.0%'
for col, w in zip('ABCDEFGHIJK', [7, 9, 8, 10, 14, 14, 11, 13, 14, 14, 14]):
    ws.column_dimensions[col].width = w
ws.cell(9, 1, '现分子与分母来自已提交的分析（表 2）；分母不受标签更正影响，因为这些患儿仍接受了该检查。'
              ).font = Font(F, size=9, italic=True)
ws.cell(10, 1, '注：此表只算检出率。实际更正后，时代模型、GEE、边际效应、配对亚组与图 2、图 3 都要重跑，由我完成。'
              ).font = Font(F, size=9, italic=True)
ws.cell(11, 1, '本表各格是公式，用 Excel 或 WPS 打开时会自动算出结果；在网页预览或部分看图工具里可能显示为空白，属正常。'
              ).font = Font(F, size=9, italic=True, color='C00000')

wb.save('/home/user/-Diagnostic-performance-of-intestinal-malrotation/待核标签清单_9例.xlsx')
print('saved')
